import requests
from dotenv import load_dotenv
import os
import pandas as pd
import json
from openpyxl import Workbook, load_workbook


# Load .env file
load_dotenv()
extracted_items = []
failed_repos = []
error_file = "error_request.xlsx"

def lazy_paginated_request(datos):
    for item in datos:
        print("--------------------ITEM-----------------",item['job'])
        headers = {
        "Authorization": "Bearer",
        "User-Agent": "MyApp/1.0",
        "Accept": "application/json",
        "Cookie": 
            """
            GCP_IAP_UID=100348192445950711918; GCP_IAP_XSRF_NONCE_WPa7-IVSEwg20IWmxpCJ5Q=1; __Host-GCP_IAP_AUTH_TOKEN_4230143DB71CD3D1=AeEuYjgI7nPMrMMwztt5rtJ6K8lZZdDMEboJRUiEi5RqQVfRerYkfRogKXCYeKGpChUAKUNZJsch62ToPmmlSoV02YjhHChZMl4tzCWSMiuDgCn6YERb45j1ba6Npi2BWu011dthKXpl2Lfak1I_mbB42T1faHLHW-6DL4R99VEZ-PkTUssygNbn_KE5HDIeOxl8ekpc4aH_ItaHQng-hKBZz4w5LSv8lbhMG1lUGurPMFH3oRp_gYO88rU2skSzC0ViPC8cFXW5BDpbC-pITD-hS9rNVeO8Ev01ri8Ygow-kNmgCG0ThsgyeeAcxSus63HM_pn6TpNtSQ2JZKS8PUeCpoJLy6hFGelptk1tdhvEReAfqc8vLZVgS-dcClX1uO4n6MCEJ4hReg4nQ0tKE3CURcHcT8K7HEWAv16-ttHZ4NQKtalzTZwBH8cyUw3fIIwbeKTWNr-rhFNT5PwXamGIgUR1N6ybvdP-x7Gp0Dkam-bhiuoQfKeu4J4CA8u_1OwCcjXr6Zp3x2LSJwV1a8oeOzwyPwfgdk90TciJ3knnUo4i3_w4rtxmfNwOo0PhY4mAqDWZidrm7lGZkGdmOyvL6t5oXP5Jljj7GB0oo4BT9CIYddRPNKhQHBIduCxlLKE3r3LYwB5tz-trYIvlSBkTc2SZMiWYIUvSXbKoQe3CVsA-BkwyCMj5I5pNNt-DHBG67wPoie2_7HWvBnEJtn74ZUxBXmGV6PTmF2fG_xR3JwG4-38zW-oyI4ohLEcczLxNlQBVZekJduBEEGzvHwS2pqLakR7ffTCfu_9r1RNwVf046428bEJWy0SnSeBOGJ4Y7Xyg3VvvABgyz0CVZQOq3x-WR4f7yFcgxVtAgxW1LnHqm_Ni2xJp4_Hwa_w-SOBf8Nb7QqiaEIB3PPLMrlIExTytFyxRNkCQrX6Oco6cR1ic-iw1yaYOEi0Q-uKNSbAxj0U3qgS1aBqPGNvlJQw20TH98xG7-RZO9PPLL0s4vFOsqTqBGhGoX8y61hll1lI6Q1bUvjscvsqPaH_6cckUVk-SVFuLvdP80K9Ky1NA_1aiQgynGhSKf3nzcBRLvU-3KZ3K-YnOmH3MwlaTn4TdnwanUKfiDp97OmYEh0yniMFY_CgOgLTvItDGiCS0vF-32jnt2IA93dqO7r00VBiIBBHpDJQX9TWoM-z-iUG1Ks3PDbEjsj_7onlsNEbuuxwTehcn4hX8l9bht4xV-lgKmmESMjAbDRzwY4MU9USFDNe1pEh0KPWsUSHFGVXy8nl3lLlPy6SCxvR5SA8DSqmw7UZh0CzXDjfffzziRFxDxa8kgUyl3LNOm25lJAtW3SX0ZCDwdZKXoydZ5mUqzP2IOkhqGP9B-dN5treLWumZc5jylMu8FSQYxd1esEOYaDKG0P7cUv3zjBhonxtjqOI1gu_CavzrkCHssOprIE0W9xapCXvjifhztQOcOub6ZJCso6wUTHxdS9_7B6zGkVNi2zAeHlWtLRF2WgKsuvhoAgOlXUts02yJWOmRA1mXrfPwT_V7Efl8V7MRz5fD7BpBQFF0WJHSkodPWg5iqkyAyR2OkNDyUXyGh8d84HBnuVEyCjwiXGbRr6wfKryRMlBuo-5wVeo6ke8J15FAPIhEV7NrKZ7lGMMCbpfHVrYkB3c8lNfid99x3YicmUFP4i9HPzo0auI4PxbgrvKEIcnvJaImYo-05OkwxZ_Lhjk2f2R8i7lr0Jf1wdvDN0k28WRxzdVV
            """
        }
        
        headers["Cookie"] = headers["Cookie"].strip()

        url = "https://bbva-lrba.appspot.com/gateway/ecs-live-02/lrba/v0/status?paginationKey=0&pageSize=1&jobName={jobName}&runId=&jobVersion=&status=&size=&priorityClassName=&namespace=&startDate=1730437200000000000&endDate=&sort=running%2Cdesc&sort=finishDate%2Cdesc&sort=startDate%2Cdesc".format(jobName=item["job"])
        
        page = 1
        while True:
            pagination = "0"
            #url = "https://bbva-lrba.appspot.com/gateway/ecs-live-02/lrba/v0/ns/co.cbtq.app-id-20452.pro/jobconfig?paginationKey=0&pageSize=200&jobName=&priorityClassName="
            response = requests.get(url, headers=headers)
            print(response.status_code)
            # print(response)
            if response.status_code == 200:
                data = response.json()
                print(data)
                if(data["result"]):
                    result = data["result"][0]
                    print(result)
                    extracted_items.append({
                            "job": result["jobName"],
                            "version": result["jobVersion"],
                            "artifact": (result.get("jobConfig") or {}).get("artifact"),
                            # "artifact": result.get("jobConfig", {}).get("artifact"),
                            "check": False
                        })
                    break
                # yield from items
                # page += 1
                # print(f"✅ Ejecutando request en consola lrba pagina {page} de {totalPages}")
                break
            else:
                print("Request failed")
                data = response.json()
                failed_repos.append([item["job"]])
                print(data)
                break


def getJob(items):
    result = []
    for item in items:
        result.append(item["jobName"])

    df = pd.DataFrame(result)
    excel_path = os.path.join("./", f'descargas_cbtq.xlsx')
    df.to_excel(excel_path, index=False, sheet_name='Archivos Descargados')
    print(f"📄 Excel generado: {excel_path}")
    return result

def initialize():
# if __name__ == "__main__":
    with open(os.path.join(os.path.dirname(__file__), "data.json"), "r", encoding="utf-8") as f:
        datos = json.load(f)

    letras_a_excluir = ('k', 'o', 'w', 'a')
    # letras_a_excluir = ('cbtq')
    # letras_a_excluir = ('a')

    resultado_filtrado = [
        item for item in datos
        if not item['job'].lower().startswith(letras_a_excluir) # Solo UUAA'S Colombia
        # if item['job'][:4].lower().startswith(letras_a_excluir)  # Solo CBTQ
        # if item['job'].lower().startswith(letras_a_excluir)  # Solo Alpha
    ]

    print(len(resultado_filtrado))
    lazy_paginated_request(resultado_filtrado)
    # print("datos--------------------------",datos)
    # print("extracted--------------------------",filter)
    try:
        with open(os.path.join(os.path.dirname(__file__), "final.json"), 'w', encoding='utf-8') as archivo:
            # La función 'dump' serializa y escribe en el archivo
            json.dump(extracted_items, archivo, ensure_ascii=False, indent=4)
        print(f"El objeto JSON se ha guardado exitosamente en ")

    except IOError as e:
        print(f"Error al escribir en el archivo: {e}")

            # Guardar errores en Excel
    if failed_repos:
        if os.path.exists(error_file):
            # Si ya existe, lo cargamos y agregamos filas
            wb = load_workbook(error_file)
            ws = wb.active
        else:
            # Crear uno nuevo con encabezados
            wb = Workbook()
            ws = wb.active
            ws.append(["Job", "Version", "Artifact", "Error"])

        for row in failed_repos:
            ws.append(row)

        wb.save(error_file)
        print(f"📊 Errores registrados en {error_file}")

    # print("Total:", len(all_results))

