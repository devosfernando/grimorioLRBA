import subprocess
import os
import re
from openpyxl import Workbook, load_workbook
import json
import git
import scripts.constants as constants

result = []

def parse_bitbucket_url(url, version):

    branch = f"tags/{version}"
    dsg = ""

    DSG_MAPPING = {
        "CO_CBGH": "CO_CBGH_APP-ID-1187023_DSG",
        "CO_CPDE": "CO_CPDE_APP-ID-62244_DSG",
        "CO_CSAN": "CO_CSAN_APP-ID-20585_DSG",
        "CO_CMOL": "CO_CMOL_APP-ID-20538_DSG",
        "CO_CPAD": "CO_CPAD_APP-ID-20553_DSG",
        "CO_CCOG": "CO_CCOG_APP-ID-20455_DSG",
        "CO_JV0D": "CO_JV0D_APP-ID-1070392_DSG",
        "CO_CUGH": "CO_CUGH_APP-ID-1182854_DSG",
        "CO_CTSU": "CO_CTSU_APP-ID-2694760_DSG",
        "CO_CBTQ": "CO_CBTQ_APP-ID-20452_DSG",
        "CO_CBGU": "CO_CBGU_APP-ID-2694715_DSG"
    }
    prefix = url[:4]
    new_url = "CO_"+ url
    # prefix = "CBGH"
    for prefix, dsg_value in DSG_MAPPING.items():
        # print(dsg_value)
        if new_url.startswith(prefix):
            dsg = dsg_value
            break # Detiene la búsqueda al encontrar la primera coincidencia
    # print(dsg)

    if(dsg):
        print("--DSG ENCONTRADO EN DICCIONARIO---")
    else:
        dsg =  url[0:4] + "_CO_DSG"

    project = dsg.lower()
    repo = url

    # https://bbva.ghe.com/platform/atau-gl-jsprk-alertreprocess-v00.git
    # git_url = f"https://bitbucket.globaldevtools.bbva.com/bitbucket/scm/{project}/{repo}.git"
    print("repo__________",{repo})
    git_url = f"https://bbva.ghe.com/platform/{repo}.git"
    return git_url, branch, url

def clone_repo(repo_urls, destination_dir="./source"):
    failed_repos = []
    error_file = os.path.join(os.path.dirname(__file__), constants.DATA_FOLDER, constants.ERROR_CLONE_XLSX)
    for repo in repo_urls:
        print(repo)
        try:
            job = repo["job"]
            version = repo["version"]
            # artifact = repo["artifact"]
            git_url,branch, repo_name = parse_bitbucket_url(job, version)

            # print(f"🧩 Clonando repo: {git_url} en la rama: {branch}")
            repo_path = os.path.join(os.path.dirname(__file__),destination_dir, repo_name)
            if not os.path.exists(repo_path):
                subprocess.check_call(["git", "clone", git_url, repo_path])
                subprocess.run(["git", "checkout", f"tags/{version}"], cwd=repo_path, check=True)
                print("✅ Clonado exitoso.")
            else:
                print(f"⚠️ Ya existe {repo_path}, se omite clonación.")
                tag_actual = obtener_tag_actual(repo_path)
                print(f"🔖 Tag actual en {repo_path}: {tag_actual}")
                print(f"✅ version {version}")
                if(tag_actual != f"Tag: {version}"):
                    subprocess.run(["git", "fetch", "origin", "--tags"], cwd=repo_path, check=True)
                    subprocess.run(["git", "checkout", f"tags/{version}"], cwd=repo_path, check=True)
                


        except Exception as e:
            error_msg = str(e)
            print(f"❌ Error al clonar {repo.get('job')}: {error_msg}")
            failed_repos.append([repo.get("job"), repo.get("version"), repo.get("artifact"), error_msg])


    # Guardar errores en Excel
    if failed_repos:
        if os.path.exists(error_file):
            # Si ya existe, lo cargamos y agregamos filas
            wb = load_workbook(error_file)
            ws = wb.active
        else:
            # Crear uno nuevo con encabezados
            wb = Workbook()
            ws = wb.active
            ws.append(["Job", "Version", "Artifact", "Error"])

        for row in failed_repos:
            ws.append(row)

        wb.save(error_file)
        print(f"📊 Errores registrados en {error_file}")
        
def obtener_tag_actual(ruta):
    try:
        # Busca el tag exacto en el que se encuentra el HEAD
        comando = ["git", "-C", ruta, "describe", "--tags", "--exact-match"]
        tag = subprocess.check_output(comando, stderr=subprocess.STDOUT).decode("utf-8").strip()
        return f"Tag: {tag}"
    except subprocess.CalledProcessError:
        # Si no hay un tag exacto, podrías estar en un commit suelto
        return "HEAD desprendido (Sin Tag exacto)"

# Ejemplo de uso:
# if __name__ == "__main__":
def initialize():   
    ruta = os.path.join(os.path.dirname(__file__), constants.DATA_FOLDER, constants.FINAL_JSON)
    print("🚀 Iniciando proceso de clonación de repositorios...",ruta) 
    with open(os.path.join(os.path.dirname(__file__), constants.DATA_FOLDER, constants.FINAL_JSON), "r", encoding="utf-8") as f:
        datos = json.load(f)
        
    print("--------------TOTAL PROYECTOS-----------------",len(datos))
    clone_repo(datos)
