import requests
from dotenv import load_dotenv
import os
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor  # <-- La clave para la velocidad
from scripts import constants
from itertools import repeat
import time
import random


load_dotenv()
extracted_items = []
failed_repos = []
error_file = os.path.join(os.path.dirname(__file__), constants.DATA_FOLDER, constants.ERROR_REQUEST_XLSX)


def procesar_un_job(item, cookie2):
    cookie_clean = cookie2.strip()
    # print(f"cookie utilizada en el proceso: {cookie_clean}...")  # Solo mostramos parte de la cookie por seguridad
    print(f"Procesando job: {item['job']} en consola de LRBA")  # Solo mostramos parte de la cookie por seguridad
    # Definimos los headers fuera para no recrearlos en cada hilo
    HEADERS = {
        "Authorization": "Bearer",
        "User-Agent": "MyApp/1.0",
        "Accept": "application/json",
        "Cookie": cookie_clean
    }
    """Función que maneja la petición individual (lo que antes estaba dentro del for)"""
    job_name = item['job']
    
    url = f"https://bbva-lrba.appspot.com/gateway/ecs-live-02/lrba/v0/status?paginationKey=0&pageSize=1&jobName={job_name}&runId=&jobVersion=&status=&size=&priorityClassName=&namespace=&startDate=1730437200000000000&endDate=&sort=running%2Cdesc&sort=finishDate%2Cdesc&sort=startDate%2Cdesc"
    
    try:
        time.sleep(random.uniform(0.1, 0.5))  # Pequeña pausa aleatoria para evitar sobrecargar el servidor
        response = requests.get(url, headers=HEADERS, timeout=10) # Añadido timeout por seguridad
  
        if response.status_code == 200:
            data = response.json()

            if data.get("result"):
                result = data["result"][0]
                # Retornamos el diccionario para agregarlo a la lista global de forma segura
                return {
                    "tipo": "exito",
                    "datos": {
                        "job": result["jobName"],
                        "version": result["jobVersion"],
                        "artifact": (result.get("jobConfig") or {}).get("artifact"),
                        "check": False
                    }
                }
        else:
            print(f"Error en la petición para {job_name}: {response.status_code}")
            return {"tipo": "error", "job": job_name, "status": response.status_code}
            
    except Exception as e:
        print(f"Excepción al procesar {job_name}: {str(e)}")
        return {"tipo": "error", "job": job_name, "error": str(e)}




def lazy_paginated_request(datos,cookie_string):
    # max_workers define cuántas peticiones se hacen al mismo tiempo. 
    # 10 es un número seguro, puedes probar con 20 si el servidor lo aguanta.
    with ThreadPoolExecutor(max_workers=10) as executor:
        # Mapeamos la función a nuestra lista de datos
        resultados_hilos = list(executor.map(procesar_un_job, datos,repeat(cookie_string)))

    # Clasificamos los resultados obtenidos de los hilos
    for res in resultados_hilos:
        if res:
            if res["tipo"] == "exito":
                extracted_items.append(res["datos"])
            else:
                failed_repos.append([res.get("job"), "N/A", "N/A", res.get("status") or res.get("error")])

# ... El resto de tus funciones (getJob, initialize) se mantienen igual ...

def initialize(cookie):
    # LIMPIEZA DE LISTAS GLOBALES (Fundamental para Streamlit)
    global extracted_items, failed_repos
    extracted_items.clear()
    failed_repos.clear()
    # (Tu lógica de carga de data.json y filtrado se mantiene igual)
    # ...
    with open(os.path.join(os.path.dirname(__file__),constants.DATA_FOLDER, constants.DATA_JSON), "r", encoding="utf-8") as f:
        datos = json.load(f)

    letras_a_excluir = ('k', 'o', 'w', 'a')
    resultado_filtrado = [item for item in datos if not item['job'].lower().startswith(letras_a_excluir)]

    print(f"Procesando {len(resultado_filtrado)} jobs en paralelo...")
    lazy_paginated_request(resultado_filtrado,cookie)
    
    # (Tu lógica de guardado de JSON y Excel de errores se mantiene igual)
    # ...
    print(f"Finalizado. Procesados: {len(extracted_items)}, Errores: {len(failed_repos)}")
    try:
        with open(os.path.join(os.path.dirname(__file__),constants.DATA_FOLDER, constants.FINAL_JSON), 'w', encoding='utf-8') as archivo:
            # La función 'dump' serializa y escribe en el archivo
            json.dump(extracted_items, archivo, ensure_ascii=False, indent=4)
        print(f"El objeto JSON se ha guardado exitosamente en {os.path.join(os.path.dirname(__file__),constants.DATA_FOLDER, constants.FINAL_JSON)}")

    except IOError as e:
        print(f"Error al escribir en el archivo: {e}")

            # Guardar errores en Excel
    if failed_repos:
        if os.path.exists(error_file):
            # Si ya existe, lo cargamos y agregamos filas
            wb = load_workbook(error_file)
            ws = wb.active
        else:
            # Crear uno nuevo con encabezados
            wb = Workbook()
            ws = wb.active
            ws.append(["Job", "Version", "Artifact", "Error"])

        for row in failed_repos:
            ws.append(row)

        wb.save(error_file)
        print(f"📊 Errores registrados en {error_file}")
        

# if __name__ == "__main__":
#     initialize()